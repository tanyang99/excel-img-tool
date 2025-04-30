import os

import pandas
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import time
import logging
import hashlib
import re
from typing import List, Dict, Set, Optional, Callable, Tuple
from PIL import Image as PILImage
from PIL import UnidentifiedImageError
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.exceptions import InvalidFileException

# Maximum file count
MAX_FILE_COUNT = 10
MAX_TOTAL_SIZE = 500 * 1024 * 1024
MAX_CONCURRENT_DOWNLOADS = 3

# 支持的图片扩展名常量
SUPPORTED_IMAGE_EXTENSIONS = ('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.svg', '.webp')


class ExcelImageEmbedder:
    def __init__(self):
        self._successfully_downloaded_urls: Set[str] = set()

    @staticmethod
    def is_image_url(value: str) -> bool:
        """
        判断单元格值是否为可能是图片地址的字符串
        :param value: 单元格的值
        :return: 如果是字符串且符合图片URL格式返回True，否则返回False
        """

        if not isinstance(value, str):
            return False
        ext_pattern = '|'.join(ext[1:] for ext in SUPPORTED_IMAGE_EXTENSIONS)  # 去掉点号
        pattern = rf'^(https?://).*\.({ext_pattern})$'
        print(value, "-------->", bool(re.match(pattern, value.lower())))
        return bool(re.match(pattern, value.lower()))

    def _download_image(self, url: str, save_path: str) -> Optional[str]:
        """
        下载图片并保存到指定路径
        :param url: 图片URL
        :param save_path: 保存路径
        :return: 保存路径如果下载成功，否则返回 None
        """
        save_dir = os.path.dirname(save_path)
        if not os.path.exists(save_dir):
            try:
                os.makedirs(save_dir, exist_ok=True)
            except OSError as e:
                logging.error(f"创建目录 {save_dir} 失败: {e}")
                return None

        if not os.access(save_dir, os.W_OK):
            logging.error(f"目录 {save_dir} 不可写。")
            return None

        if os.path.exists(save_path):
            try:
                PILImage.open(save_path).verify()
                self._successfully_downloaded_urls.add(url)
                logging.debug(f"图片 {url} 已存在于 {save_path}，跳过下载。")
                return save_path
            except UnidentifiedImageError:
                logging.warning(f"图片 {save_path} 存在但损坏，重新下载。")
                os.remove(save_path)

        try:
            session = requests.Session()
            retries = Retry(total=3, backoff_factor=1, status_forcelist=[502, 503, 504])
            session.mount('http://', HTTPAdapter(max_retries=retries))
            session.mount('https://', HTTPAdapter(max_retries=retries))

            response = session.get(url, stream=True, timeout=10)
            response.raise_for_status()

            os.makedirs(os.path.dirname(save_path), exist_ok=True)
            with open(save_path, 'wb') as file:
                for chunk in response.iter_content(chunk_size=8192):
                    file.write(chunk)

            try:
                PILImage.open(save_path).verify()
                self._successfully_downloaded_urls.add(url)
                logging.debug(f"图片 {url} 下载成功，保存到 {save_path}")
                return save_path
            except UnidentifiedImageError:
                logging.error(f"下载的图片 {url} 无效，删除文件 {save_path}")
                try:
                    os.remove(save_path)
                except OSError as e:
                    logging.error(f"删除无效图片文件 {save_path} 失败: {e}")
                return None

        except requests.exceptions.Timeout:
            logging.error(f"下载图片 {url} 时发生超时错误。")
            return None
        except requests.exceptions.HTTPError as http_err:
            logging.error(f"下载图片 {url} 时发生HTTP错误: {http_err}")
            return None
        except requests.exceptions.RequestException as req_err:
            logging.error(f"下载图片 {url} 失败: {e}")
            if os.path.exists(save_path):
                try:
                    os.remove(save_path)
                except OSError as remove_err:
                    logging.error(f"删除文件 {save_path} 失败: {remove_err}")
            return None
        except OSError as e:
            logging.error(f"下载图片 {url} 失败: {e}")
            if os.path.exists(save_path):
                try:
                    os.remove(save_path)
                except OSError as remove_err:
                    logging.error(f"删除文件 {save_path} 失败: {remove_err}")
            return None

    def _embed_image_to_cell(self, ws: Worksheet, img_path: str, row_index: int, col_index: int) -> bool:
        cell_coordinate = f'{chr(65 + col_index)}{row_index + 1}'
        try:
            if not os.path.exists(img_path):
                logging.error(f"在单元格 {cell_coordinate} 嵌入图片时出错: 图片文件不存在于 {img_path}")
                return False

            # 获取原图尺寸
            with PILImage.open(img_path) as img:
                original_width, original_height = img.size

            # 计算缩放比例，最大尺寸为 100x100，不放大
            max_size = 100
            scale = min(max_size / original_width, max_size / original_height, 1)
            new_width = int(original_width * scale)
            new_height = int(original_height * scale)

            # 创建 openpyxl Image
            img = Image(img_path)
            img.width = new_width
            img.height = new_height
            ws.add_image(img, cell_coordinate)
            return True
        except UnidentifiedImageError as e:
            logging.error(f"无法识别图片 {img_path}: {e}")
            return False
        except (OSError, ValueError) as e:
            logging.error(f"在单元格 {cell_coordinate} 嵌入图片时出错: {e}")
            return False
        except Exception as e:
            logging.error(f"嵌入图片时发生未知错误: {e}")
            return False

    @staticmethod
    def check_file_count_and_size(file_paths: List[str]) -> bool:
        """
        检查文件数量和总大小
        :param file_paths: 文件路径列表
        :return: 如果文件数量和总大小符合要求返回True，否则返回False
        """
        if len(file_paths) > MAX_FILE_COUNT:
            logging.error(f"错误: 最多只能选择 {MAX_FILE_COUNT} 个文件。")
            return False
        total_size = 0
        for path in file_paths:
            path = os.path.normcase(path)  # 规范化路径以处理 macOS 案例敏感性
            if not os.path.exists(path):
                logging.error(f"错误: 文件 {path} 未找到或无法访问。")
                return False
            try:
                total_size += os.path.getsize(path)
            except OSError as e:
                logging.error(f"获取文件 {path} 大小失败：{e}")
                return False
        if total_size > MAX_TOTAL_SIZE:
            logging.error(f"错误: 选择的文件总大小不能超过 {MAX_TOTAL_SIZE / (1024 * 1024):.2f} MB。")
            return False
        return True

    @staticmethod
    def get_file_and_sheet_info(file_paths: List[str]) -> Dict[str, List[Tuple[int, str]]]:
        """
        获取文件和工作表信息，同时返回sheet_index (0-based)
        :param file_paths: 文件路径列表
        :return: 包含文件basename和其sheet信息的字典 {file_name: [(sheet_index, sheet_name), ...]}
        """
        file_sheet_info = {}
        for file_path in file_paths:
            if not os.path.exists(file_path):
                logging.warning(f"文件 {file_path} 未找到，跳过处理。")
                continue
            try:
                excel_file = pd.ExcelFile(file_path)
                sheet_names = excel_file.sheet_names
                file_name = os.path.basename(file_path)
                sheet_info = [(index, sheet_name) for index, sheet_name in enumerate(sheet_names)]
                file_sheet_info[file_name] = sheet_info
                logging.debug(f"读取文件 {file_name} 的 sheet 信息成功。")
            except (pandas.errors.ParserError, OSError) as e:
                logging.error(f"读取文件 {file_path} 的 sheet 信息出错: {e}")
        return file_sheet_info

    def _collect_image_urls(self, wb, file_basename: str, sheets_to_process: List[int]) -> Dict[str, str]:
        """
        收集选定sheets中的图片URL
        :param wb: 工作簿对象
        :param file_basename: 文件名
        :param sheets_to_process: 需要处理的sheet索引列表
        :return: URL到保存路径的映射
        """
        urls_to_download: Set[str] = set()
        url_save_path_map: Dict[str, str] = {}
        sheet_names = wb.sheetnames

        logging.info(f"--- 收集文件 {file_basename} 中选定 sheets 的图片链接 ---")
        for sheet_index in sheets_to_process:
            if not (0 <= sheet_index < len(sheet_names)):
                logging.warning(f"文件 {file_basename}: 指定的 Sheet 索引 {sheet_index} (0-based) 不存在，跳过。")
                continue
            sheet_name = sheet_names[sheet_index]
            logging.debug(f"正在收集 Sheet: {sheet_name} (Index: {sheet_index}) 的链接...")
            ws = wb[sheet_name]
            for row_index, row in enumerate(ws.iter_rows()):
                for col_index, cell in enumerate(row):
                    if self.is_image_url(cell.value):
                        url = cell.value.strip()
                        print("1", url)
                        if url not in urls_to_download:
                            urls_to_download.add(url)
                            url_hash = hashlib.md5(url.encode('utf-8')).hexdigest()
                            ext = os.path.splitext(url.lower())[1]
                            if not ext or ext not in SUPPORTED_IMAGE_EXTENSIONS:
                                ext = '.jpg'
                            img_filename = f"{url_hash}{ext}"
                            save_path = os.path.join("downloaded_images", img_filename)
                            url_save_path_map[url] = save_path
        print("2", url_save_path_map)
        return url_save_path_map

    def _download_images(self, url_save_path_map: Dict[str, str]) -> Dict[str, Optional[str]]:
        """
        顺序下载图片
        :param url_save_path_map: URL到保存路径的映射
        :return: 下载结果映射 {url: save_path or None}
        """
        download_results: Dict[str, Optional[str]] = {}
        logging.info(f"--- 开始下载图片 ({len(url_save_path_map)} 张) ---")
        for url, save_path in url_save_path_map.items():
            save_path = self._download_image(url, save_path)
            download_results[url] = save_path
        successful_downloads = sum(1 for path in download_results.values() if path is not None)
        failed_downloads = len(url_save_path_map) - successful_downloads
        logging.info(f"--- 图片下载完成：成功 {successful_downloads} 张，失败 {failed_downloads} 张 ---")
        return download_results

    def _embed_images_to_sheets(self, wb, file_basename: str, sheets_to_process: List[int],
                                download_results: Dict[str, Optional[str]]) -> int:
        """
        将下载的图片嵌入到选定sheets
        :param wb: 工作簿对象
        :param file_basename: 文件名
        :param sheets_to_process: 需要处理的sheet索引列表
        :param download_results: 下载结果映射
        :return: 成功嵌入的图片数量
        """
        logging.info(f"--- 开始嵌入文件 {file_basename} 中选定 sheets 的图片 ---")
        successful_embeds = 0
        failed_embeds = 0
        total_attempted_embeds = 0
        sheet_names = wb.sheetnames

        for sheet_index in sheets_to_process:
            if not (0 <= sheet_index < len(sheet_names)):
                continue
            sheet_name = sheet_names[sheet_index]
            logging.debug(f"正在嵌入 Sheet: {sheet_name} (Index: {sheet_index}) 的图片...")
            ws = wb[sheet_name]
            for row_index, row in enumerate(ws.iter_rows()):
                for col_index, cell in enumerate(row):
                    if self.is_image_url(cell.value):
                        url = cell.value.strip()
                        total_attempted_embeds += 1
                        downloaded_path = download_results.get(url)
                        if downloaded_path and self._embed_image_to_cell(ws, downloaded_path, row_index, col_index):
                            successful_embeds += 1
                        else:
                            logging.error(
                                f"在单元格 {chr(65 + col_index)}{row_index + 1} 嵌入图片时出错: 图片 {url} 下载失败或嵌入失败。")
                            failed_embeds += 1

        logging.info(f"--- 图片嵌入完成：成功 {successful_embeds} 张，失败 {failed_embeds} 张，"
                     f"共检查 {total_attempted_embeds} 个包含图片链接的单元格 ---")
        return successful_embeds

    def _save_output_file(self, wb, file_basename: str) -> None:
        """
        保存修改后的Excel文件
        :param wb: 工作簿对象
        :param file_basename: 文件名
        """
        output_dir = "excel_with_images"
        os.makedirs(output_dir, exist_ok=True)
        new_file_name = f"{os.path.splitext(file_basename)[0]}_with_images.xlsx"
        new_file_path = os.path.join(output_dir, new_file_name)
        try:
            wb.save(new_file_path)
            logging.info(f"-> 已保存修改后的文件到 {new_file_path}")
        except (OSError, InvalidFileException) as e:
            logging.error(f"保存修改后的文件 {new_file_path} 时出错: {e}")

    def embed_images(self, file_paths: List[str], sheets_to_process_map: Dict[str, List[int]],
                     progress_callback: Optional[Callable[[str], None]] = None) -> None:
        """
        嵌入图片到Excel文件中
        :param file_paths: 原始文件路径列表
        :param sheets_to_process_map: 包含需要处理的工作表索引的字典
        :param progress_callback: Optional callback to report progress
        :return: None. Logs success/failure.
        """
        start_time = time.time()
        logging.info("\n \n")
        logging.info("-------------- 开始图片嵌入处理 --------------")
        logging.info(f"待处理文件: {[os.path.basename(p) for p in file_paths]}")
        logging.info(f"待处理 sheets (索引): {sheets_to_process_map}")
        if progress_callback:
            progress_callback("开始图片嵌入处理...")

        if not self.check_file_count_and_size(file_paths):
            logging.error("文件数量或大小不符合要求，终止处理。")
            if progress_callback:
                progress_callback("文件数量或大小不符合要求，终止处理。")
            return

        total_files_processed = 0
        total_successful_files = 0

        for file_path in file_paths:
            file_basename = os.path.basename(file_path)
            sheets_to_process = sheets_to_process_map.get(file_basename, [])
            if not sheets_to_process:
                logging.info(f"文件 {file_basename} 没有选定的 sheet 进行处理，跳过。")
                if progress_callback:
                    progress_callback(f"文件 {file_basename} 没有选定的 sheet 进行处理，跳过。")
                continue

            total_files_processed += 1
            logging.info(f"-> 开始处理文件: {file_basename}")
            if progress_callback:
                progress_callback(f"开始处理文件: {file_basename}")
            self._successfully_downloaded_urls.clear()

            try:
                wb = load_workbook(file_path)
                url_save_path_map = self._collect_image_urls(wb, file_basename, sheets_to_process)
                if not url_save_path_map:
                    logging.info(f"文件 {file_basename} 无图片链接，跳过。")
                    if progress_callback:
                        progress_callback(f"文件 {file_basename} 无图片链接，跳过。")
                    continue

                download_results = self._download_images(url_save_path_map)
                successful_embeds = self._embed_images_to_sheets(wb, file_basename, sheets_to_process, download_results)
                if successful_embeds > 0:
                    self._save_output_file(wb, file_basename)
                    total_successful_files += 1
                else:
                    logging.info(f"文件 {file_basename} 没有图片被处理或尝试嵌入，不生成新文件。")
                    if progress_callback:
                        progress_callback(f"文件 {file_basename} 没有图片被处理或尝试嵌入，不生成新文件。")

            except FileNotFoundError:
                logging.error(f"错误: 处理文件时 {file_path} 未找到。")
                if progress_callback:
                    progress_callback(f"错误: 处理文件时 {file_path} 未找到。")
            except (OSError, InvalidFileException) as e:
                logging.error(f"处理文件 {file_path} 时发生错误: {e}")
                if progress_callback:
                    progress_callback(f"处理文件 {file_path} 时发生错误: {str(e)}")

        logging.info("-------------- 图片嵌入处理结束 --------------")
        logging.info(f"总共处理了 {total_files_processed} 个文件，成功生成 {total_successful_files} 个带图片的输出文件。")
        logging.info(f"总处理耗时: {time.time() - start_time:.2f} 秒")
        if progress_callback:
            progress_callback(f"图片嵌入处理结束，总共处理了 {total_files_processed} 个文件，耗时: {time.time() - start_time:.2f} 秒")
